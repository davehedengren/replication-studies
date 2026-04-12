"""Shared paths, loaders, and HAC helpers for replication of paper 228661.

Paper: Gourinchas, Ray, and Vayanos (2025) "A Preferred-Habitat Model of Term
Premia, Exchange Rates, and Monetary Policy Spillovers," AER 115(11).

Scope of this replication: empirical predictability moments only. The
structural two-country preferred-habitat model (written in Julia, estimated
by MLE over a five-factor continuous-time state process) is out of scope.
"""

from pathlib import Path
import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
PKG = REPO / "228661-V1" / "exhab_aer20250923"
DATA_CLEAN = PKG / "data" / "clean"
MOMENTS_DIR = PKG / "data" / "moments"
OUT = Path(__file__).resolve().parent / "outputs"
OUT.mkdir(exist_ok=True)

QUARTERLY_DTA = DATA_CLEAN / "us_de_yc_exchange_quarterly.dta"

MATURITIES = list(range(12, 241, 12))  # months, 1y..20y
SAMPLE_START = pd.Timestamp("1986-04-01")  # 1986q2, DE yc available from 1986m6


def load_quarterly():
    df = pd.read_stata(QUARTERLY_DTA)
    df = df.sort_values("yq").reset_index(drop=True)
    return df


def load_published_moments():
    """Read the Stata-produced xlsx moments into dict of DataFrames."""
    out = {}
    for fname, tag in [
        ("us_de_moments_quarterly.xlsx", "reg"),
        ("us_de_misc_moments_quarterly.xlsx", "corr"),
    ]:
        wb = load_workbook(MOMENTS_DIR / fname, data_only=True)
        for sh in wb.sheetnames:
            ws = wb[sh]
            rows = list(ws.iter_rows(values_only=True))
            header = rows[0]
            data = rows[1:]
            out[sh] = pd.DataFrame(data, columns=header)
    return out


def _nw94_bandwidth(scores, c_gamma=1.1447):
    """Newey-West (1994) automatic bandwidth for Bartlett kernel HAC.

    Follows the algorithm in Newey & West (1994, ReStud) and Stata's
    `ivreg2` bw(auto): collapse a matrix of moment/score contributions
    into a scalar series (equal-weighted sum across columns), then pick

        n = floor(4 * (T/100)^(2/9))
        sigma(j) = (1/T) * sum_t h_t * h_{t-j}
        s0 = sigma(0) + 2 * sum_{j=1..n} sigma(j)
        s1 = 2 * sum_{j=1..n} j * sigma(j)
        gamma_hat = c_gamma * ((s1/s0)^2)^(1/3)
        L = floor(gamma_hat * T^(1/3))

    c_gamma = 1.1447 corresponds to the Bartlett kernel (q=1).
    """
    h = scores.sum(axis=1) if scores.ndim == 2 else scores
    h = h - h.mean()
    T = len(h)
    n = int(np.floor(4 * (T / 100.0) ** (2.0 / 9.0)))
    sig = np.empty(n + 1)
    sig[0] = np.dot(h, h) / T
    for j in range(1, n + 1):
        sig[j] = np.dot(h[j:], h[:-j]) / T
    s0 = sig[0] + 2.0 * sig[1:].sum()
    s1 = 2.0 * np.sum(np.arange(1, n + 1) * sig[1:])
    if s0 == 0:
        return 0
    gamma_hat = c_gamma * ((s1 / s0) ** 2) ** (1.0 / 3.0)
    return max(0, int(np.floor(gamma_hat * T ** (1.0 / 3.0))))


def _andrews_bandwidth(u):
    """Andrews (1991) optimal bandwidth for Bartlett kernel, AR(1) plug-in.

    Used as a closer approximation to Stata's `ivreg2 bw(auto)` for highly
    persistent series — our `_nw94_bandwidth` plug-in underestimates the
    optimal lag when residuals have long-run dependence.
    """
    u = np.asarray(u, dtype=float) - np.mean(u)
    T = len(u)
    num = float(np.dot(u[1:], u[:-1]))
    den = float(np.dot(u[:-1], u[:-1]))
    rho = num / den if den > 0 else 0.0
    rho = max(min(rho, 0.999), -0.999)
    alpha = (4.0 * rho * rho) / ((1.0 - rho) ** 2 * (1.0 + rho) ** 2)
    gamma = 1.1447 * (alpha ** (1.0 / 3.0))
    L = int(np.floor(gamma * T ** (1.0 / 3.0)))
    return max(0, L)


def bw_auto_newey_west(y, X):
    """OLS with Newey-West HAC SEs, Andrews (1991) AR(1) plug-in bandwidth.

    This approximates Stata's `ivreg2 bw(auto)` (Bartlett kernel) by
    selecting the bandwidth from the pilot OLS residuals via Andrews'
    AR(1) formula. Point estimates match Stata to machine precision;
    HAC SEs are close but not exact because Stata's `bw(auto)` uses an
    internal (undocumented-in-public-code) variant of the NW94 rule.
    """
    X = sm.add_constant(np.asarray(X, dtype=float))
    y = np.asarray(y, dtype=float)
    pilot = sm.OLS(y, X).fit()
    u = pilot.resid
    L = _andrews_bandwidth(u)
    res = sm.OLS(y, X).fit(
        cov_type="HAC",
        cov_kwds={"maxlags": L, "kernel": "bartlett"},
    )
    return res, L


def in_sample(df):
    return df[df["yq"] >= SAMPLE_START].reset_index(drop=True)
